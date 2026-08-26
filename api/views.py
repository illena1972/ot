# views.py
from django.db.models import Count
from rest_framework.viewsets import ModelViewSet
from .models import Department, Service, Position, Employee, ClothesItem, ClothesIssue,  \
    Stock, ClothesIssueItem
from .serializers import (
    DepartmentSerializer,
    ServiceSerializer,
    PositionSerializer,
    ClothesItemSerializer,
    ClothesIssueSerializer,
    EmployeeIssueReportSerializer,
    StockSerializer,
    EmployeeSerializer,
    ClothesIssueItemSerializer,
    OrderReportSerializer,
    OrderReportDetailSerializer,
)

from django.db.models import Sum
from django.db.models import F
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.decorators import action
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter
from datetime import timedelta
from io import BytesIO
from pathlib import Path
from urllib.parse import quote
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.conf import settings
from django.shortcuts import get_object_or_404

import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins






class DepartmentViewSet(ModelViewSet):
    queryset = (
        Department.objects
        .annotate(employee_count=Count("employee"))
        .order_by("name")
    )
    serializer_class = DepartmentSerializer


class ServiceViewSet(ModelViewSet):
    queryset = Service.objects.order_by("name")
    serializer_class = ServiceSerializer


class PositionViewSet(ModelViewSet):
    queryset = Position.objects.order_by("name")
    serializer_class = PositionSerializer




# GET /api/employees/{id}/
class EmployeeViewSet(ModelViewSet):
    queryset = Employee.objects.select_related(
        "department",
        "service",
        "position"
    ).order_by("last_name", "first_name")

    serializer_class = EmployeeSerializer

    filter_backends = [DjangoFilterBackend, SearchFilter]

    filterset_fields = ["department"]

    search_fields = [
        "last_name",
        "first_name",
        "middle_name",
    ]

    # -------------------------
    # ОТЧЕТ ПО СОТРУДНИКУ
    # GET /api/employees/{id}/report/
    # -------------------------
    @action(detail=True, methods=["get"])
    def report(self, request, pk=None):
        employee = self.get_object()

        items = ClothesIssueItem.objects.select_related(
            "issue",
            "item"
        ).filter(
            issue__employee=employee
        ).order_by("-issue__date_received")

        serializer = EmployeeIssueReportSerializer(items, many=True)

        return Response({
            "employee": {
                "id": employee.id,
                "last_name": employee.last_name,
                "first_name": employee.first_name,
                "middle_name": employee.middle_name,
                "department": employee.department.name if employee.department else "",
                "service": employee.service.name if employee.service else "",
                "position": employee.position.name if employee.position else "",
            },
            "items": serializer.data
        })


class ClothesItemViewSet(ModelViewSet):
    queryset = ClothesItem.objects.order_by("name")
    serializer_class = ClothesItemSerializer


class ClothesIssueViewSet(ModelViewSet):
    queryset = ClothesIssue.objects.select_related("employee").prefetch_related(
        "items"
    )
    serializer_class = ClothesIssueSerializer





class StockViewSet(ModelViewSet):
    queryset = Stock.objects.select_related("item").order_by(
        "item__name",
        "size",
        "height",
    )
    serializer_class = StockSerializer

    def create(self, request, *args, **kwargs):
        item = request.data.get("item")
        size = request.data.get("size")
        height = request.data.get("height")
        quantity = int(request.data.get("quantity", 0))

        stock, created = Stock.objects.get_or_create(
            item_id=item,
            size=size,
            height=height,
            defaults={
                "quantity": quantity,
            }
        )

        if not created:
            stock.quantity = F("quantity") + quantity
            stock.save()
            stock.refresh_from_db()

        serializer = self.get_serializer(stock)
        return Response(serializer.data)


@api_view(["GET"])
def stock_available(request):
    """
    Возвращает количество доступных единиц одежды на складе
    с учётом item, size и height.
    """
    item = request.GET.get("item")
    size = request.GET.get("size")
    height = request.GET.get("height")

    qs = Stock.objects.filter(item_id=item)

    if size:
        qs = qs.filter(size=size)
    if height:
        qs = qs.filter(height=height)

    total = qs.aggregate(quantity_total=Sum("quantity"))["quantity_total"] or 0
    return Response({"available": total})


class ClothesIssueItemViewSet(ModelViewSet):
    queryset = ClothesIssueItem.objects.all()
    serializer_class = ClothesIssueItemSerializer

    def destroy(self, request, *args, **kwargs):
        from django.db import transaction
        from django.db.models import F

        instance = self.get_object()

        with transaction.atomic():
            stock, _ = Stock.objects.select_for_update().get_or_create(
                item=instance.item,
                size=instance.size,
                height=instance.height,
                defaults={"quantity": 0},
            )

            stock.quantity = F("quantity") + instance.quantity
            stock.save()

            instance.delete()

        return Response(status=204)


@api_view(["DELETE"])
def write_off_issue_item(request, pk):
    item = ClothesIssueItem.objects.get(pk=pk)
    item.delete()
    return Response(status=204)




# отчет для заказа

def get_order_report_limit_date(request):
    report_date = request.GET.get("date")

    if not report_date:
        return timezone.now().date() + timedelta(days=180)

    parsed_date = parse_date(report_date)

    if not parsed_date:
        return None

    return parsed_date


def get_stock_quantity(item_id, size, height):
    return Stock.objects.filter(
        item_id=item_id,
        size=size,
        height=height,
    ).aggregate(
        quantity_total=Sum("quantity")
    )["quantity_total"] or 0


def enrich_order_report_row(row):
    stock_quantity = get_stock_quantity(
        row["item_id"],
        row["size"],
        row["height"],
    )
    total_quantity = row["total_quantity"] or 0

    return {
        "item_id": row["item_id"],
        "item_name": row["item__name"],
        "item_type": row["item__type"],
        "size": row["size"],
        "height": row["height"],
        "total_quantity": total_quantity,
        "stock_quantity": stock_quantity,
        "order_quantity": max(total_quantity - stock_quantity, 0),
    }

@api_view(["GET"])
def order_report(request):

    limit_date = get_order_report_limit_date(request)

    if not limit_date:
        return Response({"date": ["Укажите корректную дату отчета"]}, status=400)

    queryset = ClothesIssueItem.objects.filter(
        date_expire__isnull=False,
        date_expire__lte=limit_date
    )

    # ✅ фильтр по типу (ИСПРАВЛЕНО)
    item_type = request.GET.get("type")

    if item_type and item_type != "all":
        queryset = queryset.filter(
            item__type=item_type   # ← исправлено
        )

    data = queryset.values(

        "item_id",
        "item__name",
        "item__type",   # ← исправлено
        "size",
        "height",

    ).annotate(

        total_quantity=Sum("quantity")

    ).order_by("item__name")

    result = [enrich_order_report_row(row) for row in data]

    serializer = OrderReportSerializer(result, many=True)

    return Response(serializer.data)

# детализация отчета для заказа

@api_view(["GET"])
def order_report_detail(request):

    item_id = request.GET.get("item_id")
    size = request.GET.get("size")
    height = request.GET.get("height")

    limit_date = get_order_report_limit_date(request)

    if not limit_date:
        return Response({"date": ["Укажите корректную дату отчета"]}, status=400)

    queryset = ClothesIssueItem.objects.filter(
        item_id=item_id,
        date_expire__isnull=False,
        date_expire__lte=limit_date
    ).select_related(
        "issue",
        "issue__employee",
        "item"
    )

    # ✅ фильтр size
    if size not in [None, "", "null"]:
        queryset = queryset.filter(size=size)
    else:
        queryset = queryset.filter(size__isnull=True)

    # ✅ фильтр height
    if height not in [None, "", "null"]:
        queryset = queryset.filter(height=height)
    else:
        queryset = queryset.filter(height__isnull=True)

    # ✅ сортировка
    queryset = queryset.order_by(
        "date_expire",
        "issue__employee__last_name"
    )

    serializer = OrderReportDetailSerializer(
        queryset,
        many=True
    )

    return Response(serializer.data)




@api_view(["GET"])
def order_report_export(request):
    """
    Экспорт отчёта для заказа спецодежды в Excel.
    Берём все выданные позиции с оставшимся сроком <= 6 месяцев,
    группируем по item/size/height и суммируем количество.
    Можно фильтровать по типу одежды через GET-параметр type.
    """

    limit_date = get_order_report_limit_date(request)

    if not limit_date:
        return Response({"date": ["Укажите корректную дату отчета"]}, status=400)

    # базовый queryset
    queryset = ClothesIssueItem.objects.filter(
        date_expire__isnull=False,
        date_expire__lte=limit_date
    )

    # фильтр по типу одежды
    item_type = request.GET.get("type")
    if item_type and item_type != "all":
        queryset = queryset.filter(item__type=item_type)

    # группировка по item/size/height
    data = queryset.values(
        "item_id",
        "item__name",
        "item__type",
        "size",
        "height",
    ).annotate(
        total_quantity=Sum("quantity")
    ).order_by("item__name")

    # создаём Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт для заказа"

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.3,
        right=0.3,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2,
    )

    thin_border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    summary_fill = PatternFill("solid", fgColor="EFF6FF")

    ws.append(["Отчёт для заказа спецодежды"])
    ws.append([f"Сформировано на дату: {limit_date.strftime('%d.%m.%Y')}"])
    ws.append([
        "Наименование",
        "Размер",
        "Рост",
        "Требуется заменить",
        "Есть на складе",
        "К заказу",
        "ФИО сотрудника",
        "Кол-во",
        "Дата выдачи",
        "Дата окончания",
    ])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(italic=True, color="4B5563")

    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def get_detail_queryset(row):
        detail_queryset = ClothesIssueItem.objects.filter(
            item_id=row["item_id"],
            date_expire__isnull=False,
            date_expire__lte=limit_date,
        ).select_related(
            "issue",
            "issue__employee",
            "item",
        )

        if row["size"] is None:
            detail_queryset = detail_queryset.filter(size__isnull=True)
        else:
            detail_queryset = detail_queryset.filter(size=row["size"])

        if row["height"] is None:
            detail_queryset = detail_queryset.filter(height__isnull=True)
        else:
            detail_queryset = detail_queryset.filter(height=row["height"])

        return detail_queryset.order_by(
            "date_expire",
            "issue__employee__last_name",
            "issue__employee__first_name",
        )

    for row in data:
        report_row = enrich_order_report_row(row)

        ws.append([
            report_row["item_name"],
            report_row["size"] or "",
            report_row["height"] or "",
            report_row["total_quantity"],
            report_row["stock_quantity"],
            report_row["order_quantity"],
            "",
            "",
            "",
            "",
        ])

        summary_row = ws.max_row

        for cell in ws[summary_row]:
            cell.font = Font(bold=True)
            cell.fill = summary_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        detail_start_row = summary_row + 1

        for issue_item in get_detail_queryset(row):
            ws.append([
                "",
                "",
                "",
                "",
                "",
                "",
                str(issue_item.issue.employee),
                issue_item.quantity,
                issue_item.issue.date_received,
                issue_item.date_expire,
            ])

            detail_row = ws.max_row
            ws.row_dimensions[detail_row].outlineLevel = 1
            ws.row_dimensions[detail_row].hidden = True

            for cell in ws[detail_row]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            ws.cell(row=detail_row, column=9).number_format = "DD.MM.YYYY"
            ws.cell(row=detail_row, column=10).number_format = "DD.MM.YYYY"

        if ws.max_row >= detail_start_row:
            ws.row_dimensions[summary_row].collapsed = True

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 36
    ws.column_dimensions["H"].width = 9
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:J{ws.max_row}"
    ws.print_title_rows = "1:3"

    # формируем ответ
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="order_report.xlsx"'
    wb.save(response)

    return response


@api_view(["GET"])
def employee_card_export(request):
    """Creates a personal PPE issue card for the selected employee."""
    employee_id = request.GET.get("employee_id")
    if not employee_id:
        return Response(
            {"employee_id": ["Выберите сотрудника для формирования карточки."]},
            status=400,
        )

    employee = get_object_or_404(
        Employee.objects.select_related("department", "position"), pk=employee_id
    )
    issue_items = ClothesIssueItem.objects.filter(
        issue__employee=employee
    ).select_related("issue", "item").order_by(
        "issue__date_received", "item__name", "pk"
    )

    template_path = (
        Path(settings.BASE_DIR) / "api" / "report_templates" / "employee_card_template.xlsx"
    )
    wb = openpyxl.load_workbook(template_path)
    ws = wb["ЛК СИЗ"]

    sex_labels = {"M": "Мужской", "F": "Женский"}
    ws["M3"] = employee.last_name
    ws["H4"] = employee.first_name
    ws["AJ4"] = employee.middle_name or ""
    ws["CD3"] = sex_labels.get(employee.sex, "")
    ws["CD4"] = employee.height or ""
    ws["AE6"] = employee.department.name if employee.department else ""
    ws["CD6"] = employee.clothes_size or ""
    ws["AA7"] = employee.position.name if employee.position else ""
    ws["CD7"] = employee.shoe_size or ""

    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= 13:
            ws.unmerge_cells(str(merged_range))
    for row in ws.iter_rows(min_row=13, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None
    if ws.max_row >= 16:
        ws.delete_rows(16, ws.max_row - 15)

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_aligned = Alignment(horizontal="left", vertical="top", wrap_text=True)
    columns = [
        (1, 22, "Наименование СИЗ"),
        (23, 40, "Модель, марка, артикул, класс защиты СИЗ"),
        (41, 48, "Дата"),
        (49, 54, "Количество"),
        (55, 61, "Лично/дозатор"),
        (62, 70, "Подпись получившего СИЗ"),
        (71, 78, "Дата"),
        (79, 84, "Количество"),
        (85, 93, "Подпись сдавшего СИЗ"),
        (94, 101, "Акт списания (дата, номер)"),
    ]

    ws.merge_cells(start_row=13, start_column=1, end_row=15, end_column=22)
    ws.merge_cells(start_row=13, start_column=23, end_row=15, end_column=40)
    ws.merge_cells(start_row=13, start_column=41, end_row=13, end_column=70)
    ws.merge_cells(start_row=13, start_column=71, end_row=13, end_column=101)
    ws.cell(13, 1, "Наименование СИЗ")
    ws.cell(13, 23, "Модель, марка, артикул, класс защиты СИЗ")
    ws.cell(13, 41, "Выдано")
    ws.cell(13, 71, "Возвращено")
    for start, end, title in columns[2:]:
        ws.merge_cells(start_row=14, start_column=start, end_row=15, end_column=end)
        ws.cell(14, start, title)

    for row_number in range(13, 16):
        for column_number in range(1, 102):
            cell = ws.cell(row_number, column_number)
            cell.border = thin_border
            cell.fill = header_fill
            cell.font = Font(bold=True, size=8)
            cell.alignment = centered

    for row_number, issue_item in enumerate(issue_items, start=16):
        values = [
            issue_item.item.name,
            "",
            issue_item.issue.date_received.strftime("%d.%m.%Y"),
            issue_item.quantity,
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        for (start, end, _), value in zip(columns, values):
            ws.merge_cells(start_row=row_number, start_column=start, end_row=row_number, end_column=end)
            cell = ws.cell(row_number, start, value)
            cell.border = thin_border
            cell.alignment = left_aligned if start in (1, 23) else centered
            cell.font = Font(size=8)
            for column_number in range(start, end + 1):
                ws.cell(row_number, column_number).border = thin_border
        ws.row_dimensions[row_number].height = 30

    if not issue_items:
        ws.merge_cells(start_row=16, start_column=1, end_row=16, end_column=101)
        cell = ws.cell(16, 1, "Выдачи СИЗ отсутствуют")
        cell.alignment = centered
        cell.font = Font(italic=True, size=9, color="6B7280")
        for column_number in range(1, 102):
            ws.cell(16, column_number).border = thin_border

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.35, bottom=0.35, header=0.1, footer=0.1)
    ws.print_area = f"A1:CW{ws.max_row}"
    ws.print_title_rows = "1:15"

    output = BytesIO()
    wb.save(output)
    filename = f"Карточка_СИЗ_{employee.last_name}_{employee.first_name}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=employee_card.xlsx; filename*=UTF-8''{quote(filename)}"
    )
    return response
